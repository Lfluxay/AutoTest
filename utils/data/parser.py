import yaml
import json
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Union, Tuple
from functools import lru_cache
from utils.logging.logger import logger
from utils.data.template_parser import template_parser


class DataParser:
    """数据解析器 - 支持YAML和Excel格式的用例数据解析"""
    
    @staticmethod
    @lru_cache(maxsize=256)
    def _parse_yaml_cached(key: Tuple[str, float]) -> Dict[str, Any]:
        file_path_str, _mtime = key
        file_path = Path(file_path_str)
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)

    @staticmethod
    def parse_yaml(file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        解析YAML文件（带mtime缓存）
        """
        file_path = Path(file_path)
        try:
            mtime = file_path.stat().st_mtime
            data = DataParser._parse_yaml_cached((str(file_path), mtime))
            logger.debug(f"成功解析YAML文件: {file_path}")
            return data
        except FileNotFoundError:
            logger.error(f"YAML文件不存在: {file_path}")
            raise
        except yaml.YAMLError as e:
            logger.error(f"YAML文件格式错误 {file_path}: {e}")
            raise
    
    @staticmethod
    @lru_cache(maxsize=64)
    def _parse_excel_cached(key: Tuple[str, float, str]) -> List[Dict[str, Any]]:
        file_path_str, _mtime, sheet_name = key
        file_path = Path(file_path_str)
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        # 获取表头
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value)
            else:
                break
        # 解析数据行
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    if headers[i] in ['headers', 'data', 'params', 'assertions', 'extract']:
                        try:
                            if isinstance(value, str) and value.strip():
                                row_data[headers[i]] = json.loads(value)
                            else:
                                row_data[headers[i]] = {}
                        except (json.JSONDecodeError, TypeError):
                            row_data[headers[i]] = value
                    else:
                        row_data[headers[i]] = value
            if row_data:
                data.append(row_data)
        return data

    @staticmethod
    def parse_excel(file_path: Union[str, Path], sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        解析Excel文件（带mtime缓存）
        """
        file_path = Path(file_path)
        try:
            mtime = file_path.stat().st_mtime
            data = DataParser._parse_excel_cached((str(file_path), mtime, sheet_name or ''))
            logger.debug(f"成功解析Excel文件: {file_path}, 工作表: {sheet_name or 'default'}, 数据条数: {len(data)}")
            return data
        except FileNotFoundError:
            logger.error(f"Excel文件不存在: {file_path}")
            raise
        except Exception as e:
            logger.error(f"Excel文件解析错误 {file_path}: {e}")
            raise
    
    @staticmethod
    def parse_json(json_str: str) -> Any:
        """
        解析JSON字符串
        
        Args:
            json_str: JSON字符串
            
        Returns:
            解析后的数据
        """
        try:
            return json.loads(json_str)
        except json.JSONDecodeError as e:
            logger.error(f"JSON格式错误: {e}")
            raise
    
    @staticmethod
    def convert_excel_to_yaml(excel_path: Union[str, Path], yaml_path: Union[str, Path], sheet_name: str = None):
        """
        将Excel文件转换为YAML文件
        
        Args:
            excel_path: Excel文件路径
            yaml_path: 输出的YAML文件路径  
            sheet_name: 工作表名称
        """
        data = DataParser.parse_excel(excel_path, sheet_name)
        
        # 转换为YAML格式
        yaml_data = {
            "test_info": {
                "title": f"从Excel转换的测试用例: {Path(excel_path).stem}",
                "description": f"从 {excel_path} 转换而来",
                "tags": ["converted"]
            },
            "test_cases": data
        }
        
        yaml_path = Path(yaml_path)
        with open(yaml_path, 'w', encoding='utf-8') as f:
            yaml.dump(yaml_data, f, allow_unicode=True, indent=2, default_flow_style=False)
        
        logger.info(f"Excel转YAML完成: {excel_path} -> {yaml_path}")

    @staticmethod
    def convert_excel_to_template_yaml(excel_path: Union[str, Path], yaml_path: Union[str, Path],
                                     template_name: str, sheet_name: str = None):
        """
        将Excel文件转换为模板化YAML文件

        Args:
            excel_path: Excel文件路径
            yaml_path: 输出的YAML文件路径
            template_name: 使用的模板名称
            sheet_name: 工作表名称
        """
        data = DataParser.parse_excel(excel_path, sheet_name)

        # 检查是否包含template列
        if data and 'template' in data[0]:
            # Excel中已指定模板，按模板分组
            template_cases = {}
            for row in data:
                template = row.get('template', 'default_template')
                if template not in template_cases:
                    template_cases[template] = []

                # 移除template列，其余作为测试数据
                test_data = {k: v for k, v in row.items() if k != 'template'}
                template_cases[template].append(test_data)

            # 生成模板化YAML
            yaml_data = {
                "test_info": {
                    "title": f"从Excel转换的模板化测试用例: {Path(excel_path).stem}",
                    "description": f"从 {excel_path} 转换而来",
                    "tags": ["converted", "template"]
                },
                "test_cases": []
            }

            for template, test_data_list in template_cases.items():
                case_config = {
                    "case_name": f"Excel转换-{template}",
                    "template": template,
                    "test_data": test_data_list,
                    "tags": ["excel_converted"]
                }
                yaml_data["test_cases"].append(case_config)
        else:
            # 使用指定的模板名称
            yaml_data = {
                "test_info": {
                    "title": f"从Excel转换的模板化测试用例: {Path(excel_path).stem}",
                    "description": f"从 {excel_path} 转换而来，使用模板 {template_name}",
                    "tags": ["converted", "template"]
                },
                "test_cases": [
                    {
                        "case_name": f"Excel转换-{template_name}",
                        "template": template_name,
                        "test_data": data,
                        "tags": ["excel_converted"]
                    }
                ]
            }

        yaml_path = Path(yaml_path)
        with open(yaml_path, 'w', encoding='utf-8') as f:
            yaml.dump(yaml_data, f, allow_unicode=True, indent=2, default_flow_style=False)

        logger.info(f"Excel转模板化YAML完成: {excel_path} -> {yaml_path}")

    @staticmethod
    def create_excel_template(excel_path: Union[str, Path], template_type: str = "api"):
        """
        创建Excel模板文件

        Args:
            excel_path: Excel文件路径
            template_type: 模板类型 (api/web/workflow)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active

        if template_type == "api":
            # API测试用例模板
            headers = [
                "case_name", "template", "username", "password",
                "expected_status", "expected_code", "expected_message", "tags"
            ]
            example_data = [
                ["管理员登录", "templates/api_templates.yaml#login_template",
                 "admin", "123456", "200", "0", "登录成功", "positive,smoke"],
                ["用户登录", "templates/api_templates.yaml#login_template",
                 "user001", "user123", "200", "0", "登录成功", "positive"],
                ["登录失败", "templates/api_templates.yaml#login_template",
                 "admin", "wrong", "400", "1002", "密码错误", "negative"]
            ]
        elif template_type == "web":
            # Web测试用例模板
            headers = [
                "case_name", "template", "username", "password",
                "login_should_succeed", "expected_url_keyword", "tags"
            ]
            example_data = [
                ["Web管理员登录", "templates/web_templates.yaml#login_page_template",
                 "admin", "admin123", "true", "dashboard", "positive,smoke"],
                ["Web用户登录", "templates/web_templates.yaml#login_page_template",
                 "user001", "user123", "true", "home", "positive"],
                ["Web登录失败", "templates/web_templates.yaml#login_page_template",
                 "admin", "wrong", "false", "login", "negative"]
            ]
        elif template_type == "workflow":
            # 工作流测试用例模板
            headers = [
                "case_name", "template", "register_username", "register_email",
                "register_password", "update_enabled", "expected_workflow_success", "tags"
            ]
            example_data = [
                ["用户完整流程", "templates/api_templates.yaml#user_workflow_template",
                 "flow_user_001", "flow001@example.com", "Flow123", "true", "true", "workflow,positive"],
                ["用户注册登录流程", "templates/api_templates.yaml#user_workflow_template",
                 "flow_user_002", "flow002@example.com", "Flow456", "false", "true", "workflow,positive"]
            ]

        # 设置表头
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # 添加示例数据
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 保存文件
        excel_path = Path(excel_path)
        wb.save(excel_path)

        logger.info(f"Excel模板文件创建完成: {excel_path} ({template_type}类型)")
    
    @staticmethod
    def validate_case_format(case_data: Dict[str, Any], case_type: str = "api") -> bool:
        """
        验证用例数据格式
        
        Args:
            case_data: 用例数据
            case_type: 用例类型 (api/web)
            
        Returns:
            验证结果
        """
        try:
            # 基本字段验证
            if 'case_name' not in case_data:
                logger.error("用例缺少必填字段: case_name")
                return False
            
            if case_type == "api":
                # API用例验证
                if 'request' not in case_data:
                    logger.error(f"API用例 {case_data['case_name']} 缺少request字段")
                    return False
                
                request = case_data['request']
                if 'method' not in request or 'url' not in request:
                    logger.error(f"API用例 {case_data['case_name']} request缺少method或url字段")
                    return False
                    
            elif case_type == "web":
                # Web用例验证
                if 'steps' not in case_data:
                    logger.error(f"Web用例 {case_data['case_name']} 缺少steps字段")
                    return False
                
                steps = case_data['steps']
                if not isinstance(steps, list) or not steps:
                    logger.error(f"Web用例 {case_data['case_name']} steps必须是非空列表")
                    return False
            
            logger.debug(f"用例格式验证通过: {case_data['case_name']}")
            return True
        except Exception as e:
            logger.error(f"用例格式验证失败: {e}")
            return False

    def get_test_cases(self, file_path: str, case_filter: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """
        获取测试用例列表，支持传统用例和模板化用例

        Args:
            file_path: 测试用例文件路径
            case_filter: 用例过滤条件

        Returns:
            测试用例列表
        """
        test_data = self.parse_yaml(file_path)

        if not test_data:
            return []

        # 检查是否是模板化用例
        if self._is_template_based_case(test_data):
            return self._process_template_cases(test_data, case_filter, file_path)

        # 传统用例处理
        test_cases = test_data.get('test_cases', [])

        if case_filter:
            test_cases = self._filter_test_cases(test_cases, case_filter)

        return test_cases

    def _is_template_based_case(self, test_data: Dict[str, Any]) -> bool:
        """
        检查是否是模板化用例

        Args:
            test_data: 测试数据

        Returns:
            是否是模板化用例
        """
        test_cases = test_data.get('test_cases', [])

        # 检查是否有用例使用了template字段
        for case in test_cases:
            if isinstance(case, dict) and ('template' in case or 'dataset' in case):
                return True

        return False

    def _process_template_cases(self, test_data: Dict[str, Any], case_filter: Dict[str, Any] = None,
                              file_path: str = None) -> List[Dict[str, Any]]:
        """
        处理模板化用例

        Args:
            test_data: 测试数据
            case_filter: 用例过滤条件
            file_path: 当前文件路径，用于解析文件内模板

        Returns:
            生成的测试用例列表
        """
        all_generated_cases = []
        test_cases = test_data.get('test_cases', [])

        # 设置文件内模板的上下文
        if file_path:
            template_parser.set_context_file(file_path)

        for case_config in test_cases:
            if isinstance(case_config, dict):
                if 'template' in case_config:
                    # 模板化用例
                    try:
                        generated_cases = template_parser.generate_test_cases(case_config, file_path)
                        all_generated_cases.extend(generated_cases)
                        logger.info(f"模板用例生成成功: {case_config.get('case_name', 'Unknown')}, 生成 {len(generated_cases)} 个用例")
                    except Exception as e:
                        logger.error(f"模板用例生成失败: {case_config.get('case_name', 'Unknown')}, 错误: {e}")
                else:
                    # 传统用例
                    all_generated_cases.append(case_config)

        # 应用过滤条件
        if case_filter:
            all_generated_cases = self._filter_test_cases(all_generated_cases, case_filter)

        return all_generated_cases

    def _filter_test_cases(self, test_cases: List[Dict[str, Any]], case_filter: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        过滤测试用例

        Args:
            test_cases: 测试用例列表
            case_filter: 过滤条件

        Returns:
            过滤后的测试用例列表
        """
        if not case_filter:
            return test_cases

        filtered_cases = []

        for case in test_cases:
            # 按标签过滤
            if 'tags' in case_filter:
                required_tags = case_filter['tags']
                case_tags = case.get('tags', [])

                if isinstance(required_tags, str):
                    required_tags = [required_tags]

                # 检查是否有交集
                if not any(tag in case_tags for tag in required_tags):
                    continue

            # 按严重级别过滤
            if 'severity' in case_filter:
                required_severity = case_filter['severity']
                case_severity = case.get('severity', 'normal')

                if case_severity != required_severity:
                    continue

            # 按用例名称过滤
            if 'case_name' in case_filter:
                required_name = case_filter['case_name']
                case_name = case.get('case_name', '')

                if required_name not in case_name:
                    continue

            filtered_cases.append(case)

        return filtered_cases