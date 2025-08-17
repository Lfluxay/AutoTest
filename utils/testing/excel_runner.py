"""
Excel直接测试运行器
支持Excel文件直接执行测试，每个Sheet作为一个测试模块
"""
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from utils.logging.logger import logger
from utils.data.template_parser import template_parser
from utils.core.exceptions import AutoTestException
from utils.core.base.test_case import TestCase, TestSuite, TestCaseType, TestCaseSource, TestRequest, TestAssertion, TestExtraction


class ExcelTestRunner:
    """Excel测试运行器"""
    
    def __init__(self, project_root: Optional[str] = None):
        """
        初始化Excel测试运行器
        
        Args:
            project_root: 项目根目录
        """
        self.project_root = Path(project_root) if project_root else Path(__file__).parent.parent
        
    def parse_excel_file(self, excel_path: Union[str, Path]) -> Dict[str, List[Dict[str, Any]]]:
        """
        解析Excel文件，每个Sheet作为一个测试模块
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            模块字典，key为模块名(sheet名)，value为测试用例列表
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise AutoTestException(f"Excel文件不存在: {excel_path}", "EXCEL_FILE_NOT_FOUND")
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            modules = {}
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"解析Excel模块: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # 跳过空sheet
                if sheet.max_row <= 1:
                    logger.warning(f"跳过空Sheet: {sheet_name}")
                    continue
                
                test_cases = self.parse_sheet(sheet, sheet_name)
                if test_cases:
                    modules[sheet_name] = test_cases
                    logger.info(f"模块 {sheet_name} 解析完成: {len(test_cases)} 个用例")
                else:
                    logger.warning(f"模块 {sheet_name} 没有有效用例")
            
            workbook.close()
            return modules
            
        except Exception as e:
            raise AutoTestException(f"解析Excel文件失败: {excel_path}, 错误: {e}", "EXCEL_PARSE_ERROR")
    
    def parse_sheet(self, sheet, sheet_name: str) -> List[Dict[str, Any]]:
        """
        解析单个Sheet
        
        Args:
            sheet: openpyxl worksheet对象
            sheet_name: Sheet名称
            
        Returns:
            测试用例列表
        """
        # 获取表头
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        
        if not headers:
            logger.warning(f"Sheet {sheet_name} 没有有效表头")
            return []
        
        logger.debug(f"Sheet {sheet_name} 表头: {headers}")
        
        # 检查用例类型
        has_template = 'template' in headers
        has_method = 'method' in headers
        has_type = 'type' in headers
        has_action = 'action' in headers
        has_url = 'url' in headers
        
        test_cases = []
        
        # 从第二行开始解析数据
        for row_num in range(2, sheet.max_row + 1):
            row_data = {}
            
            # 读取行数据
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                value = cell.value
                
                # 处理空值
                if value is None:
                    value = ""
                else:
                    value = str(value).strip()
                
                row_data[header] = value
            
            # 跳过空行或无用例名称的行
            if not row_data.get('case_name'):
                continue
            
            # 检查是否启用
            enabled = row_data.get('enabled', 'Y')
            if enabled.upper() in ['N', 'NO', 'FALSE', '0']:
                logger.debug(f"跳过禁用用例: {row_data.get('case_name')}")
                continue
            
            # 解析用例
            try:
                if has_type:
                    # 混合模式：根据type字段决定解析方式
                    case_type = row_data.get('type', '').lower()
                    if case_type == 'template':
                        test_case = self.parse_template_case(row_data, sheet_name)
                    elif case_type == 'custom':
                        test_case = self.parse_custom_case(row_data, sheet_name)
                    else:
                        # 自动判断
                        if row_data.get('template'):
                            test_case = self.parse_template_case(row_data, sheet_name)
                        else:
                            test_case = self.parse_custom_case(row_data, sheet_name)
                elif has_template:
                    # 模板模式
                    test_case = self.parse_template_case(row_data, sheet_name)
                elif has_method or has_action or has_url:
                    # 自定义模式（API或Web）
                    test_case = self.parse_custom_case(row_data, sheet_name)
                else:
                    logger.warning(f"无法识别用例类型: {row_data.get('case_name')}")
                    continue
                
                if test_case:
                    test_cases.append(test_case)
                    
            except Exception as e:
                logger.error(f"解析用例失败: {row_data.get('case_name')}, 错误: {e}")
                continue
        
        return test_cases
    
    def parse_template_case(self, row_data: Dict[str, Any], module_name: str) -> Optional[Dict[str, Any]]:
        """
        解析模板化用例
        
        Args:
            row_data: 行数据
            module_name: 模块名称
            
        Returns:
            解析后的测试用例
        """
        template_path = row_data.get('template')
        if not template_path:
            logger.warning(f"模板用例缺少template字段: {row_data.get('case_name')}")
            return None
        
        try:
            # 加载模板
            template = template_parser.load_template(template_path)
            
            # 准备变量字典
            variables = {}
            for key, value in row_data.items():
                if key not in ['case_name', 'template', 'type', 'enabled', 'tags']:
                    variables[key] = value
            
            # 替换模板变量
            test_case = template_parser.substitute_variables(template, variables)
            
            # 添加用例元信息
            test_case['case_name'] = row_data.get('case_name')
            test_case['description'] = template.get('description', '')
            test_case['module'] = module_name
            test_case['source'] = 'excel_template'
            
            # 处理标签
            tags = row_data.get('tags', '').split(',') if row_data.get('tags') else []
            tags = [tag.strip() for tag in tags if tag.strip()]
            tags.append(module_name)  # 添加模块名作为标签
            test_case['tags'] = tags
            
            # 过滤条件项目
            if 'assertions' in test_case:
                test_case['assertions'] = template_parser.filter_by_condition(test_case['assertions'], variables)
            if 'extract' in test_case:
                test_case['extract'] = template_parser.filter_by_condition(test_case['extract'], variables)
            if 'steps' in test_case:
                test_case['steps'] = template_parser.filter_by_condition(test_case['steps'], variables)
            
            return test_case
            
        except Exception as e:
            logger.error(f"解析模板用例失败: {row_data.get('case_name')}, 错误: {e}")
            return None
    
    def parse_custom_case(self, row_data: Dict[str, Any], module_name: str) -> Optional[Dict[str, Any]]:
        """
        解析自定义用例
        
        Args:
            row_data: 行数据
            module_name: 模块名称
            
        Returns:
            解析后的测试用例
        """
        try:
            test_case = {
                'case_name': row_data.get('case_name'),
                'description': row_data.get('description', ''),
                'module': module_name,
                'source': 'excel_custom'
            }
            
            # 构建请求信息
            request = {}
            
            # HTTP方法
            method = row_data.get('method', '').upper()
            if method:
                request['method'] = method
            
            # URL
            url = row_data.get('url', '')
            if url:
                request['url'] = url
            
            # 请求头
            headers = row_data.get('headers', '')
            if headers:
                try:
                    import json
                    request['headers'] = json.loads(headers)
                except:
                    # 简单格式：key1:value1,key2:value2
                    header_dict = {}
                    for item in headers.split(','):
                        if ':' in item:
                            key, value = item.split(':', 1)
                            header_dict[key.strip()] = value.strip()
                    if header_dict:
                        request['headers'] = header_dict
            
            # 请求数据
            data = row_data.get('data', '')
            if data:
                try:
                    import json
                    request['data'] = json.loads(data)
                except:
                    # 简单格式处理
                    request['data'] = data
            
            # 请求参数
            params = row_data.get('params', '')
            if params:
                try:
                    import json
                    request['params'] = json.loads(params)
                except:
                    # 简单格式：key1=value1&key2=value2
                    param_dict = {}
                    for item in params.split('&'):
                        if '=' in item:
                            key, value = item.split('=', 1)
                            param_dict[key.strip()] = value.strip()
                    if param_dict:
                        request['params'] = param_dict
            
            if request:
                test_case['request'] = request
            
            # 断言
            assertions_str = row_data.get('assertions', '')
            if assertions_str:
                try:
                    import json
                    test_case['assertions'] = json.loads(assertions_str)
                except:
                    # 简单断言：status_code=200,json_path=$.code:0
                    assertions = []
                    for item in assertions_str.split(','):
                        if '=' in item:
                            assertion_type, expected = item.split('=', 1)
                            assertion = {
                                'type': assertion_type.strip(),
                                'expected': expected.strip()
                            }
                            # 处理json_path特殊格式
                            if ':' in expected and assertion_type.strip() == 'json_path':
                                path, value = expected.split(':', 1)
                                assertion['path'] = path.strip()
                                assertion['expected'] = value.strip()
                            assertions.append(assertion)
                    if assertions:
                        test_case['assertions'] = assertions
            
            # 提取
            extract_str = row_data.get('extract', '')
            if extract_str:
                try:
                    import json
                    test_case['extract'] = json.loads(extract_str)
                except:
                    # 简单提取：token=$.data.token,user_id=$.data.user.id
                    extracts = []
                    for item in extract_str.split(','):
                        if '=' in item:
                            name, path = item.split('=', 1)
                            extract = {
                                'name': name.strip(),
                                'type': 'json_path',
                                'path': path.strip()
                            }
                            extracts.append(extract)
                    if extracts:
                        test_case['extract'] = extracts
            
            # 处理标签
            tags = row_data.get('tags', '').split(',') if row_data.get('tags') else []
            tags = [tag.strip() for tag in tags if tag.strip()]
            tags.append(module_name)  # 添加模块名作为标签
            test_case['tags'] = tags
            
            return test_case
            
        except Exception as e:
            logger.error(f"解析自定义用例失败: {row_data.get('case_name')}, 错误: {e}")
            return None
    
    def filter_modules(self, modules: Dict[str, List[Dict[str, Any]]], 
                      module_filter: Optional[List[str]] = None,
                      tag_filter: Optional[List[str]] = None) -> Dict[str, List[Dict[str, Any]]]:
        """
        过滤模块和用例
        
        Args:
            modules: 模块字典
            module_filter: 模块过滤列表
            tag_filter: 标签过滤列表
            
        Returns:
            过滤后的模块字典
        """
        filtered_modules = {}
        
        for module_name, test_cases in modules.items():
            # 模块过滤
            if module_filter and module_name not in module_filter:
                continue
            
            # 标签过滤
            if tag_filter:
                filtered_cases = []
                for case in test_cases:
                    case_tags = case.get('tags', [])
                    if any(tag in case_tags for tag in tag_filter):
                        filtered_cases.append(case)
                if filtered_cases:
                    filtered_modules[module_name] = filtered_cases
            else:
                filtered_modules[module_name] = test_cases
        
        return filtered_modules
    
    def generate_yaml_from_excel(self, excel_path: Union[str, Path], 
                                output_dir: Union[str, Path]) -> List[str]:
        """
        从Excel生成YAML文件（每个模块一个文件）
        
        Args:
            excel_path: Excel文件路径
            output_dir: 输出目录
            
        Returns:
            生成的YAML文件路径列表
        """
        modules = self.parse_excel_file(excel_path)
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        generated_files = []
        
        for module_name, test_cases in modules.items():
            yaml_data = {
                'test_info': {
                    'title': f'{module_name}模块测试',
                    'description': f'从Excel文件 {Path(excel_path).name} 的 {module_name} Sheet生成',
                    'tags': [module_name, 'excel_generated'],
                    'module': module_name
                },
                'test_cases': test_cases
            }
            
            yaml_file = output_dir / f"{module_name}_cases.yaml"
            
            import yaml
            with open(yaml_file, 'w', encoding='utf-8') as f:
                yaml.dump(yaml_data, f, allow_unicode=True, indent=2, default_flow_style=False)
            
            generated_files.append(str(yaml_file))
            logger.info(f"生成YAML文件: {yaml_file}")
        
        return generated_files

    def parse_to_test_suite(self, excel_path: Union[str, Path],
                           module_filter: Optional[List[str]] = None,
                           tag_filter: Optional[List[str]] = None) -> TestSuite:
        """
        解析Excel文件为TestSuite对象

        Args:
            excel_path: Excel文件路径
            module_filter: 模块过滤列表
            tag_filter: 标签过滤列表

        Returns:
            TestSuite对象
        """
        excel_path = Path(excel_path)

        # 解析Excel文件
        modules = self.parse_excel_file(excel_path)

        # 应用过滤条件
        filtered_modules = self.filter_modules(modules, module_filter, tag_filter)

        # 创建测试套件
        test_suite = TestSuite(
            title=f"Excel测试套件: {excel_path.stem}",
            description=f"从Excel文件 {excel_path.name} 解析生成",
            tags=["excel", "auto_generated"],
            module="excel_import",
            version="1.0"
        )

        # 转换为TestCase对象
        for module_name, test_cases in filtered_modules.items():
            for case_dict in test_cases:
                test_case = self._dict_to_test_case(case_dict, module_name)
                test_suite.add_test_case(test_case)

        logger.info(f"Excel解析完成: {len(test_suite.test_cases)} 个测试用例")
        return test_suite

    def _dict_to_test_case(self, case_dict: Dict[str, Any], module_name: str) -> TestCase:
        """
        将字典转换为TestCase对象

        Args:
            case_dict: 用例字典
            module_name: 模块名称

        Returns:
            TestCase对象
        """
        # 智能识别测试类型
        case_type = self._detect_test_type(case_dict)

        # 基本信息
        test_case = TestCase(
            case_name=case_dict.get("case_name", "Unknown"),
            description=case_dict.get("description", ""),
            case_type=case_type,
            source=TestCaseSource.EXCEL,
            module=module_name,
            tags=case_dict.get("tags", []),
            severity=case_dict.get("severity", "normal"),
            enabled=True
        )

        # 根据测试类型设置相应的字段
        if case_type == TestCaseType.API:
            self._set_api_fields(test_case, case_dict)
        elif case_type == TestCaseType.WEB:
            self._set_web_fields(test_case, case_dict)

        # 断言
        if "assertions" in case_dict:
            test_case.assertions = []
            for assertion_data in case_dict["assertions"]:
                assertion = TestAssertion(
                    type=assertion_data.get("type", ""),
                    expected=assertion_data.get("expected"),
                    operator=assertion_data.get("operator", "=="),
                    path=assertion_data.get("path"),
                    message=assertion_data.get("message"),
                    condition=assertion_data.get("condition"),
                    max_time=assertion_data.get("max_time"),
                    attribute=assertion_data.get("attribute"),
                    locator=assertion_data.get("locator")
                )
                test_case.assertions.append(assertion)

        # 提取
        if "extract" in case_dict:
            test_case.extract = []
            for extract_data in case_dict["extract"]:
                extraction = TestExtraction(
                    name=extract_data.get("name", ""),
                    type=extract_data.get("type", ""),
                    path=extract_data.get("path"),
                    locator=extract_data.get("locator"),
                    condition=extract_data.get("condition")
                )
                test_case.extract.append(extraction)

        # 模板信息
        if case_dict.get("source") == "excel_template":
            test_case.template_path = case_dict.get("template")
            test_case.test_data = case_dict.get("test_data")

        return test_case

    def _detect_test_type(self, case_dict: Dict[str, Any]) -> TestCaseType:
        """
        智能识别测试类型

        Args:
            case_dict: 用例字典

        Returns:
            TestCaseType: 测试类型
        """
        # 1. 显式指定测试类型
        test_type = case_dict.get("test_type", "").lower()
        if test_type == "web":
            return TestCaseType.WEB
        elif test_type == "api":
            return TestCaseType.API

        # 2. 根据字段特征智能识别
        # Web测试特征字段（权重更高的放前面）
        web_indicators = [
            ("action", 4),  # action字段是Web测试的强特征
            ("locator", 4), ("selector", 4), ("element_visible", 4),
            ("steps", 3), ("navigate", 3), ("click", 3), ("input", 3),
            ("wait_for_element", 3), ("element_text", 3), ("element_attribute", 3),
            ("page_url", 2), ("browser", 2), ("driver", 2)
        ]

        # API测试特征字段（权重更高的放前面）
        api_indicators = [
            ("method", 4),  # HTTP方法是API测试的强特征
            ("json_path", 4), ("status_code", 4),
            ("headers", 3), ("endpoint", 3), ("request", 3), ("response", 3),
            ("data", 2), ("params", 2)  # data和params可能在Web中也有
        ]

        # 检查字段名和值
        case_str = str(case_dict).lower()
        field_names = [key.lower() for key in case_dict.keys()]

        web_score = 0
        api_score = 0

        # 计算Web特征分数
        for indicator, weight in web_indicators:
            if indicator in field_names:
                web_score += weight * 3  # 字段名匹配权重更高
            elif indicator in case_str:
                web_score += weight

        # 计算API特征分数
        for indicator, weight in api_indicators:
            if indicator in field_names:
                api_score += weight * 3  # 字段名匹配权重更高
            elif indicator in case_str:
                api_score += weight

        # 特殊处理：如果有断言内容，检查断言类型
        assertions_str = str(case_dict.get("assertions", "")).lower()
        if "element_" in assertions_str or "url_contains" in assertions_str:
            web_score += 5  # Web断言强特征
        elif "json_path" in assertions_str or "status_code" in assertions_str:
            api_score += 5  # API断言强特征

        # 3. 根据模板路径判断
        template = case_dict.get("template", "")
        if template:
            if "web" in template.lower():
                return TestCaseType.WEB
            elif "api" in template.lower():
                return TestCaseType.API

        # 4. 根据URL特征判断
        url = case_dict.get("url", "")
        if url:
            if url.startswith(("http://", "https://")) and not url.startswith(("http://localhost", "https://localhost")):
                # 外部URL更可能是API测试
                return TestCaseType.API
            elif "localhost" in url or any(port in url for port in [":3000", ":8080", ":9000"]):
                # 本地服务器更可能是Web测试
                return TestCaseType.WEB

        # 5. 根据评分决定
        if web_score > api_score:
            return TestCaseType.WEB
        else:
            return TestCaseType.API  # 默认为API测试

    def _set_api_fields(self, test_case: TestCase, case_dict: Dict[str, Any]):
        """设置API测试字段"""
        # 请求信息
        if "request" in case_dict:
            req_data = case_dict["request"]
            test_case.request = TestRequest(
                method=req_data.get("method", "GET"),
                url=req_data.get("url", ""),
                headers=req_data.get("headers"),
                data=req_data.get("data"),
                params=req_data.get("params"),
                files=req_data.get("files"),
                timeout=req_data.get("timeout")
            )
        else:
            # 从Excel列直接构建请求
            test_case.request = TestRequest(
                method=case_dict.get("method", "GET"),
                url=case_dict.get("url", ""),
                headers=case_dict.get("headers"),
                data=case_dict.get("data"),
                params=case_dict.get("params"),
                files=case_dict.get("files"),
                timeout=case_dict.get("timeout")
            )

    def _set_web_fields(self, test_case: TestCase, case_dict: Dict[str, Any]):
        """设置Web测试字段"""
        from utils.core.base.test_case import TestStep

        # Web步骤
        if "steps" in case_dict:
            test_case.steps = []
            for step_data in case_dict["steps"]:
                step = TestStep(
                    action=step_data.get("action", ""),
                    params=step_data.get("params", {}),
                    condition=step_data.get("condition")
                )
                test_case.steps.append(step)
        else:
            # 从Excel列构建简单的Web步骤
            steps = []

            # 获取URL（可能在request字段中）
            url = case_dict.get("url", "").strip()
            if not url and "request" in case_dict:
                url = case_dict["request"].get("url", "").strip()

            action = case_dict.get("action", "").strip()

            if url and not action:
                # 只有URL，默认为导航操作
                steps.append(TestStep(
                    action="navigate",
                    params={"url": url},
                    condition=None
                ))
            elif action:
                # 有明确的action
                params = {}

                # 构建参数
                if case_dict.get("locator"):
                    params["locator"] = case_dict.get("locator")
                if case_dict.get("value"):
                    params["value"] = case_dict.get("value")
                if case_dict.get("text"):
                    params["text"] = case_dict.get("text")
                if url and action == "navigate":
                    params["url"] = url

                steps.append(TestStep(
                    action=action,
                    params=params,
                    condition=case_dict.get("condition")
                ))
            elif url:
                # 有URL但action为空，默认导航
                steps.append(TestStep(
                    action="navigate",
                    params={"url": url},
                    condition=None
                ))
            else:
                # 没有URL和action，但可能有其他操作信息
                # 根据断言类型推断操作
                if case_dict.get("assertions"):
                    for assertion in case_dict["assertions"]:
                        if isinstance(assertion, dict):
                            assertion_type = assertion.get("type", "")
                            if assertion_type == "element_value" and assertion.get("locator"):
                                # 元素值断言，可能是input操作
                                steps.append(TestStep(
                                    action="input",
                                    params={
                                        "locator": assertion.get("locator"),
                                        "value": assertion.get("expected", "")
                                    },
                                    condition=None
                                ))
                                break

            test_case.steps = steps if steps else None


# 全局Excel测试运行器实例
excel_runner = ExcelTestRunner()
